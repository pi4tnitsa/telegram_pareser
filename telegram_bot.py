import asyncio
import datetime
import json
import logging
import os
import re
import sqlite3
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import ParseMode, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton
from telethon import TelegramClient
from telethon.tl.functions.channels import JoinChannelRequest
from telethon.tl.functions.messages import GetHistoryRequest
from telethon.errors import ChannelPrivateError

# Import configuration
from config import api_id, api_hash, BOT_TOKEN, ADMIN_IDS

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    filename='bot_logs.log', filemode='a')
logger = logging.getLogger(__name__)

# Create temp directory if it doesn't exist
if not os.path.exists('temp'):
    os.makedirs('temp')

# Initialize bot and dispatcher
bot = Bot(token=BOT_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)

# Initialize Telethon client
client = TelegramClient('bot_session', api_id, api_hash)

# Define states for conversation handlers
class ExportStates(StatesGroup):
    select_data_type = State()
    select_period = State()
    custom_period_start = State()
    custom_period_end = State()
    select_format = State()

class SourceStates(StatesGroup):
    add_source = State()
    delete_source = State()
    edit_source = State()
    confirm_delete = State()

class KeywordStates(StatesGroup):
    add_keyword = State()
    delete_keyword = State()
    confirm_delete = State()

class SearchStates(StatesGroup):
    enter_query = State()
    select_period = State()
    custom_period_start = State()
    custom_period_end = State()

# Helper functions for database operations
def init_db():
    """Initialize database and create tables if they don't exist"""
    conn = sqlite3.connect('telegram_content.db')
    cursor = conn.cursor()
    
    # Create tables if they don't exist
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS posts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT,
        channel_name TEXT,
        content TEXT,
        message_id INTEGER
    )
    ''')
    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS comments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT,
        channel_name TEXT,
        post_content TEXT,
        comment_text TEXT,
        user_id INTEGER,
        username TEXT,
        sentiment TEXT
    )
    ''')
    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT,
        source TEXT,
        content TEXT,
        user_id INTEGER,
        username TEXT,
        media_type TEXT
    )
    ''')
    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS monitored_sources (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE,
        type TEXT,
        date_added TEXT,
        is_active INTEGER DEFAULT 1
    )
    ''')
    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS keywords (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        word TEXT UNIQUE,
        date_added TEXT
    )
    ''')
    
    conn.commit()
    conn.close()

def add_source(source_name, source_type):
    """Add a new source to monitor"""
    conn = sqlite3.connect('telegram_content.db')
    cursor = conn.cursor()
    
    try:
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute(
            "INSERT INTO monitored_sources (name, type, date_added) VALUES (?, ?, ?)",
            (source_name, source_type, current_date)
        )
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()

def get_sources():
    """Get all monitored sources"""
    conn = sqlite3.connect('telegram_content.db')
    cursor = conn.cursor()
    
    cursor.execute("SELECT name, type FROM monitored_sources WHERE is_active = 1")
    sources = cursor.fetchall()
    
    conn.close()
    return sources

def delete_source(source_name):
    """Delete a source from the monitored list"""
    conn = sqlite3.connect('telegram_content.db')
    cursor = conn.cursor()
    
    cursor.execute("DELETE FROM monitored_sources WHERE name = ?", (source_name,))
    conn.commit()
    
    conn.close()

def add_keyword(keyword):
    """Add a new keyword to monitor"""
    conn = sqlite3.connect('telegram_content.db')
    cursor = conn.cursor()
    
    try:
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute(
            "INSERT INTO keywords (word, date_added) VALUES (?, ?)",
            (keyword, current_date)
        )
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()

def get_keywords():
    """Get all monitored keywords"""
    conn = sqlite3.connect('telegram_content.db')
    cursor = conn.cursor()
    
    cursor.execute("SELECT word FROM keywords")
    keywords = [row[0] for row in cursor.fetchall()]
    
    conn.close()
    return keywords

def delete_keyword(keyword):
    """Delete a keyword from the monitored list"""
    conn = sqlite3.connect('telegram_content.db')
    cursor = conn.cursor()
    
    cursor.execute("DELETE FROM keywords WHERE word = ?", (keyword,))
    conn.commit()
    
    conn.close()

def get_period_dates(period):
    """Get start and end dates based on the selected period"""
    end_date = datetime.now()
    
    if period == "week":
        start_date = end_date - timedelta(days=7)
    elif period == "month":
        start_date = end_date - timedelta(days=30)
    elif period == "three_months":
        start_date = end_date - timedelta(days=90)
    else:  # All time
        start_date = datetime(2000, 1, 1)  # A date far in the past
    
    return start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")

def export_data_to_excel(data_type, start_date, end_date):
    """Export data to Excel file"""
    conn = sqlite3.connect('telegram_content.db')
    cursor = conn.cursor()
    
    wb = openpyxl.Workbook()
    
    # Format dates for SQL query
    start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
    end_date_obj = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)  # Include the end date
    
    start_date_str = start_date_obj.strftime("%Y-%m-%d")
    end_date_str = end_date_obj.strftime("%Y-%m-%d")
    
    if data_type == "posts" or data_type == "all":
        # Export posts
        ws_posts = wb.active
        ws_posts.title = "Posts"
        
        # Add headers
        headers = ["Date", "Channel", "Content"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_posts.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Get posts data
        cursor.execute(
            "SELECT date, channel_name, content FROM posts WHERE date BETWEEN ? AND ?",
            (start_date_str, end_date_str)
        )
        posts = cursor.fetchall()
        
        # Add data to worksheet
        for row_num, post in enumerate(posts, 2):
            for col_num, value in enumerate(post, 1):
                ws_posts.cell(row=row_num, column=col_num).value = value
    
    if data_type == "comments" or data_type == "all":
        # Export comments
        if data_type == "comments":
            ws_comments = wb.active
            ws_comments.title = "Comments"
        else:
            ws_comments = wb.create_sheet("Comments")
        
        # Add headers
        headers = ["Date", "Channel", "Post Content", "Comment", "User ID", "Username", "Sentiment"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_comments.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Get comments data
        cursor.execute(
            "SELECT date, channel_name, post_content, comment_text, user_id, username, sentiment FROM comments WHERE date BETWEEN ? AND ?",
            (start_date_str, end_date_str)
        )
        comments = cursor.fetchall()
        
        # Add data to worksheet
        for row_num, comment in enumerate(comments, 2):
            for col_num, value in enumerate(comment, 1):
                ws_comments.cell(row=row_num, column=col_num).value = value
    
    if data_type == "messages" or data_type == "all":
        # Export messages
        if data_type == "messages":
            ws_messages = wb.active
            ws_messages.title = "Messages"
        else:
            ws_messages = wb.create_sheet("Messages")
        
        # Add headers
        headers = ["Date", "Source", "Content", "User ID", "Username", "Media Type"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_messages.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Get messages data
        cursor.execute(
            "SELECT date, source, content, user_id, username, media_type FROM messages WHERE date BETWEEN ? AND ?",
            (start_date_str, end_date_str)
        )
        messages = cursor.fetchall()
        
        # Add data to worksheet
        for row_num, message in enumerate(messages, 2):
            for col_num, value in enumerate(message, 1):
                ws_messages.cell(row=row_num, column=col_num).value = value
    
    # Adjust column widths
    for sheet in wb:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)  # Cap at 50 to avoid too wide columns
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    filename = f"temp/export_{data_type}_{start_date}_to_{end_date}.xlsx"
    wb.save(filename)
    conn.close()
    
    return filename

def export_data_to_json(data_type, start_date, end_date):
    """Export data to JSON file"""
    conn = sqlite3.connect('telegram_content.db')
    cursor = conn.cursor()
    
    # Format dates for SQL query
    start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
    end_date_obj = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)  # Include the end date
    
    start_date_str = start_date_obj.strftime("%Y-%m-%d")
    end_date_str = end_date_obj.strftime("%Y-%m-%d")
    
    data = {}
    
    if data_type == "posts" or data_type == "all":
        # Export posts
        cursor.execute(
            "SELECT date, channel_name, content FROM posts WHERE date BETWEEN ? AND ?",
            (start_date_str, end_date_str)
        )
        posts = cursor.fetchall()
        
        posts_data = []
        for post in posts:
            posts_data.append({
                "date": post[0],
                "channel": post[1],
                "content": post[2]
            })
        
        data["posts"] = posts_data
    
    if data_type == "comments" or data_type == "all":
        # Export comments
        cursor.execute(
            "SELECT date, channel_name, post_content, comment_text, user_id, username, sentiment FROM comments WHERE date BETWEEN ? AND ?",
            (start_date_str, end_date_str)
        )
        comments = cursor.fetchall()
        
        comments_data = []
        for comment in comments:
            comments_data.append({
                "date": comment[0],
                "channel": comment[1],
                "post_content": comment[2],
                "comment": comment[3],
                "user_id": comment[4],
                "username": comment[5],
                "sentiment": comment[6]
            })
        
        data["comments"] = comments_data
    
    if data_type == "messages" or data_type == "all":
        # Export messages
        cursor.execute(
            "SELECT date, source, content, user_id, username, media_type FROM messages WHERE date BETWEEN ? AND ?",
            (start_date_str, end_date_str)
        )
        messages = cursor.fetchall()
        
        messages_data = []
        for message in messages:
            messages_data.append({
                "date": message[0],
                "source": message[1],
                "content": message[2],
                "user_id": message[3],
                "username": message[4],
                "media_type": message[5]
            })
        
        data["messages"] = messages_data
    
    # Save to JSON file
    filename = f"temp/export_{data_type}_{start_date}_to_{end_date}.json"
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    
    conn.close()
    return filename

def search_content(query, start_date, end_date):
    """Search content based on query and period"""
    conn = sqlite3.connect('telegram_content.db')
    cursor = conn.cursor()
    
    # Format dates for SQL query
    start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
    end_date_obj = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)  # Include the end date
    
    start_date_str = start_date_obj.strftime("%Y-%m-%d")
    end_date_str = end_date_obj.strftime("%Y-%m-%d")
    
    results = []
    
    # Search in posts
    cursor.execute(
        "SELECT date, channel_name, content, 'post' as type FROM posts WHERE content LIKE ? AND date BETWEEN ? AND ?",
        (f"%{query}%", start_date_str, end_date_str)
    )
    posts = cursor.fetchall()
    results.extend(posts)
    
    # Search in comments
    cursor.execute(
        "SELECT date, channel_name, comment_text, 'comment' as type FROM comments WHERE comment_text LIKE ? AND date BETWEEN ? AND ?",
        (f"%{query}%", start_date_str, end_date_str)
    )
    comments = cursor.fetchall()
    results.extend(comments)
    
    # Search in messages
    cursor.execute(
        "SELECT date, source, content, 'message' as type FROM messages WHERE content LIKE ? AND date BETWEEN ? AND ?",
        (f"%{query}%", start_date_str, end_date_str)
    )
    messages = cursor.fetchall()
    results.extend(messages)
    
    conn.close()
    return results

def analyze_sentiment(text):
    """Simple sentiment analysis based on keywords"""
    positive_words = ['хорошо', 'отлично', 'супер', 'класс', 'радость', 'счастье', 'великолепно', 'прекрасно']
    negative_words = ['плохо', 'ужасно', 'отстой', 'проблема', 'неудача', 'грустно', 'разочарован', 'жаль']
    
    text = text.lower()
    
    positive_count = sum(1 for word in positive_words if word in text)
    negative_count = sum(1 for word in negative_words if word in text)
    
    if positive_count > negative_count:
        return "positive"
    elif negative_count > positive_count:
        return "negative"
    else:
        return "neutral"

def get_statistics():
    """Get general statistics"""
    conn = sqlite3.connect('telegram_content.db')
    cursor = conn.cursor()
    
    # Total counts
    cursor.execute("SELECT COUNT(*) FROM posts")
    posts_count = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM comments")
    comments_count = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM messages")
    messages_count = cursor.fetchone()[0]
    
    # Top 5 channels by post count
    cursor.execute(
        "SELECT channel_name, COUNT(*) as count FROM posts GROUP BY channel_name ORDER BY count DESC LIMIT 5"
    )
    top_channels = cursor.fetchall()
    
    # Posts by day of week
    cursor.execute(
        "SELECT strftime('%w', date) as day_of_week, COUNT(*) as count FROM posts GROUP BY day_of_week ORDER BY day_of_week"
    )
    posts_by_day = cursor.fetchall()
    
    # Comment sentiment distribution
    cursor.execute(
        "SELECT sentiment, COUNT(*) as count FROM comments GROUP BY sentiment"
    )
    sentiment_distribution = cursor.fetchall()
    
    # Media type distribution
    cursor.execute(
        "SELECT media_type, COUNT(*) as count FROM messages WHERE media_type IS NOT NULL GROUP BY media_type"
    )
    media_distribution = cursor.fetchall()
    
    conn.close()
    
    # Create charts
    # Day of week activity chart
    days = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    counts = [0] * 7
    for day, count in posts_by_day:
        counts[int(day)] = count
    
    plt.figure(figsize=(10, 6))
    plt.bar(days, counts, color='skyblue')
    plt.title('Activity by Day of Week')
    plt.xlabel('Day')
    plt.ylabel('Number of Posts')
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.savefig('temp/day_activity_chart.png')
    plt.close()
    
    # Sentiment distribution chart
    sentiments = []
    sentiment_counts = []
    for sentiment, count in sentiment_distribution:
        sentiments.append(sentiment)
        sentiment_counts.append(count)
    
    plt.figure(figsize=(8, 8))
    colors = ['green', 'red', 'gray']
    plt.pie(sentiment_counts, labels=sentiments, autopct='%1.1f%%', startangle=140, colors=colors)
    plt.axis('equal')
    plt.title('Comment Sentiment Distribution')
    plt.savefig('temp/sentiment_chart.png')
    plt.close()
    
    # Media type distribution chart
    media_types = []
    media_counts = []
    for media_type, count in media_distribution:
        media_types.append(media_type if media_type else "Text")
        media_counts.append(count)
    
    plt.figure(figsize=(10, 6))
    plt.bar(media_types, media_counts, color='lightgreen')
    plt.title('Media Type Distribution')
    plt.xlabel('Media Type')
    plt.ylabel('Count')
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig('temp/media_chart.png')
    plt.close()
    
    return {
        "posts_count": posts_count,
        "comments_count": comments_count,
        "messages_count": messages_count,
        "top_channels": top_channels,
        "day_activity_chart": "temp/day_activity_chart.png",
        "sentiment_chart": "temp/sentiment_chart.png",
        "media_chart": "temp/media_chart.png"
    }

async def check_keywords_in_content(content, source_name, content_type, content_date):
    """Check if content contains any monitored keywords and notify admins"""
    keywords = get_keywords()
    
    # Convert content to string in case it's not
    if content is None:
        return
    
    content_str = str(content).lower()
    
    found_keywords = [keyword for keyword in keywords if keyword.lower() in content_str]
    
    if found_keywords:
        # Create notification message
        notification = f"🔍 *Обнаружены ключевые слова:* {', '.join(found_keywords)}\n\n"
        notification += f"📂 *Тип контента:* {content_type}\n"
        notification += f"📢 *Источник:* {source_name}\n"
        notification += f"📅 *Дата:* {content_date}\n\n"
        notification += f"💬 *Содержание:*\n{content[:200]}..."
        
        # Send notification to all admins
        for admin_id in ADMIN_IDS:
            try:
                await bot.send_message(admin_id, notification, parse_mode=ParseMode.MARKDOWN)
            except Exception as e:
                logger.error(f"Failed to send notification to admin {admin_id}: {e}")

async def collect_channel_content():
    """Collect content from monitored sources"""
    sources = get_sources()
    
    conn = sqlite3.connect('telegram_content.db')
    cursor = conn.cursor()
    
    for source_name, source_type in sources:
        try:
            # Join the channel/group if not joined already
            try:
                entity = await client.get_entity(source_name)
                if hasattr(entity, 'megagroup') or hasattr(entity, 'channel'):
                    await client(JoinChannelRequest(entity))
            except ChannelPrivateError:
                logger.error(f"Cannot join private channel/group: {source_name}")
                continue
            except Exception as e:
                logger.error(f"Error joining channel/group {source_name}: {e}")
                continue
            
            # Get recent messages
            messages = await client(GetHistoryRequest(
                peer=source_name,
                limit=50,
                offset_date=None,
                offset_id=0,
                max_id=0,
                min_id=0,
                add_offset=0,
                hash=0
            ))
            
            for message in messages.messages:
                message_date = message.date.strftime("%Y-%m-%d %H:%M:%S")
                message_content = message.message
                
                if not message_content:
                    continue
                
                if source_type == "channel":
                    # Add post to database
                    cursor.execute(
                        "INSERT INTO posts (date, channel_name, content, message_id) VALUES (?, ?, ?, ?)",
                        (message_date, source_name, message_content, message.id)
                    )
                    conn.commit()
                    
                    # Check if post contains keywords
                    await check_keywords_in_content(message_content, source_name, "post", message_date)
                    
                    # Get comments if available
                    try:
                        comments = await client.get_messages(
                            entity=source_name,
                            reply_to=message.id,
                            limit=100
                        )
                        
                        for comment in comments:
                            if not comment.message:
                                continue
                                
                            comment_date = comment.date.strftime("%Y-%m-%d %H:%M:%S")
                            comment_text = comment.message
                            user_id = comment.from_id.user_id if comment.from_id else None
                            username = None
                            
                            if user_id:
                                try:
                                    user = await client.get_entity(user_id)
                                    username = user.username or f"{user.first_name} {user.last_name if user.last_name else ''}"
                                except:
                                    pass
                            
                            sentiment = analyze_sentiment(comment_text)
                            
                            # Add comment to database
                            cursor.execute(
                                "INSERT INTO comments (date, channel_name, post_content, comment_text, user_id, username, sentiment) VALUES (?, ?, ?, ?, ?, ?, ?)",
                                (comment_date, source_name, message_content, comment_text, user_id, username, sentiment)
                            )
                            conn.commit()
                            
                            # Check if comment contains keywords
                            await check_keywords_in_content(comment_text, source_name, "comment", comment_date)
                    except Exception as e:
                        logger.error(f"Error getting comments for {source_name}, message {message.id}: {e}")
                else:  # Group
                    # Determine media type
                    media_type = None
                    if message.media:
                        if hasattr(message.media, 'photo'):
                            media_type = "photo"
                        elif hasattr(message.media, 'document'):
                            if hasattr(message.media.document, 'mime_type'):
                                if 'video' in message.media.document.mime_type:
                                    media_type = "video"
                                elif 'audio' in message.media.document.mime_type:
                                    media_type = "audio"
                                else:
                                    media_type = "document"
                    
                    user_id = message.from_id.user_id if message.from_id else None
                    username = None
                    
                    if user_id:
                        try:
                            user = await client.get_entity(user_id)
                            username = user.username or f"{user.first_name} {user.last_name if user.last_name else ''}"
                        except:
                            pass
                    
                    # Add message to database
                    cursor.execute(
                        "INSERT INTO messages (date, source, content, user_id, username, media_type) VALUES (?, ?, ?, ?, ?, ?)",
                        (message_date, source_name, message_content, user_id, username, media_type)
                    )
                    conn.commit()
                    
                    # Check if message contains keywords
                    await check_keywords_in_content(message_content, source_name, "message", message_date)
                    
        except Exception as e:
            logger.error(f"Error collecting content from {source_name}: {e}")
    
    conn.close()

# Command handlers
@dp.message_handler(commands=['start', 'help'])
async def send_welcome(message: types.Message):
    """Send welcome message and show main menu"""
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(KeyboardButton("📊 Статистика"))
    keyboard.add(KeyboardButton("🔍 Поиск контента"))
    keyboard.add(KeyboardButton("📤 Экспорт данных"))
    keyboard.add(KeyboardButton("📋 Управление источниками"))
    keyboard.add(KeyboardButton("🔑 Ключевые слова"))
    
    welcome_message = (
        "👋 Добро пожаловать в бот для мониторинга Telegram-контента!\n\n"
        "Этот бот собирает данные из каналов и групп Telegram, предоставляет статистику и позволяет экспортировать данные.\n\n"
        "Воспользуйтесь меню ниже для навигации по функциям бота:"
    )
    
    await message.answer(welcome_message, reply_markup=keyboard)

@dp.message_handler(lambda message: message.text == "📤 Экспорт данных")
async def export_data_command(message: types.Message):
    """Start export data flow"""
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("Посты", callback_data="export_posts"))
    keyboard.add(InlineKeyboardButton("Комментарии", callback_data="export_comments"))
    keyboard.add(InlineKeyboardButton("Сообщения из групп", callback_data="export_messages"))
    keyboard.add(InlineKeyboardButton("Все данные", callback_data="export_all"))
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_main"))
    
    await message.answer("Выберите тип данных для экспорта:", reply_markup=keyboard)
    await ExportStates.select_data_type.set()

@dp.callback_query_handler(lambda c: c.data.startswith('export_'), state=ExportStates.select_data_type)
async def process_export_type(callback_query: types.CallbackQuery, state: FSMContext):
    """Process selected data type for export"""
    await callback_query.answer()
    
    data_type = callback_query.data.split('_')[1]
    await state.update_data(data_type=data_type)
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("Неделя", callback_data="period_week"))
    keyboard.add(InlineKeyboardButton("Месяц", callback_data="period_month"))
    keyboard.add(InlineKeyboardButton("3 месяца", callback_data="period_three_months"))
    keyboard.add(InlineKeyboardButton("Все время", callback_data="period_all"))
# Continue from where the code left off

@dp.callback_query_handler(lambda c: c.data.startswith('period_'), state=ExportStates.select_period)
async def process_export_period(callback_query: types.CallbackQuery, state: FSMContext):
    """Process selected period for export"""
    await callback_query.answer()
    
    period = callback_query.data.split('_')[1]
    
    if period == "custom":
        await callback_query.message.edit_text("Введите начальную дату в формате ГГГГ-ММ-ДД:")
        await ExportStates.custom_period_start.set()
    else:
        await state.update_data(period=period)
        
        # Get start and end dates based on period
        start_date, end_date = get_period_dates(period)
        await state.update_data(start_date=start_date, end_date=end_date)
        
        # Ask for export format
        keyboard = InlineKeyboardMarkup()
        keyboard.add(InlineKeyboardButton("Excel", callback_data="format_excel"))
        keyboard.add(InlineKeyboardButton("JSON", callback_data="format_json"))
        keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_period"))
        
        await callback_query.message.edit_text(
            f"Выбран период: {period}. Выберите формат экспорта:",
            reply_markup=keyboard
        )
        await ExportStates.select_format.set()

@dp.message_handler(state=ExportStates.custom_period_start)
async def process_custom_start_date(message: types.Message, state: FSMContext):
    """Process custom start date"""
    try:
        start_date = datetime.strptime(message.text, "%Y-%m-%d").strftime("%Y-%m-%d")
        await state.update_data(start_date=start_date)
        
        await message.answer("Введите конечную дату в формате ГГГГ-ММ-ДД:")
        await ExportStates.custom_period_end.set()
    except ValueError:
        await message.answer("Неверный формат даты. Пожалуйста, используйте формат ГГГГ-ММ-ДД:")

@dp.message_handler(state=ExportStates.custom_period_end)
async def process_custom_end_date(message: types.Message, state: FSMContext):
    """Process custom end date"""
    try:
        end_date = datetime.strptime(message.text, "%Y-%m-%d").strftime("%Y-%m-%d")
        await state.update_data(end_date=end_date)
        
        # Ask for export format
        keyboard = InlineKeyboardMarkup()
        keyboard.add(InlineKeyboardButton("Excel", callback_data="format_excel"))
        keyboard.add(InlineKeyboardButton("JSON", callback_data="format_json"))
        keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_period"))
        
        data = await state.get_data()
        
        await message.answer(
            f"Выбран период: с {data['start_date']} по {end_date}. Выберите формат экспорта:",
            reply_markup=keyboard
        )
        await ExportStates.select_format.set()
    except ValueError:
        await message.answer("Неверный формат даты. Пожалуйста, используйте формат ГГГГ-ММ-ДД:")

@dp.callback_query_handler(lambda c: c.data.startswith('format_'), state=ExportStates.select_format)
async def process_export_format(callback_query: types.CallbackQuery, state: FSMContext):
    """Process selected export format and generate export file"""
    await callback_query.answer()
    
    export_format = callback_query.data.split('_')[1]
    data = await state.get_data()
    
    data_type = data.get('data_type')
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    
    await callback_query.message.edit_text("⏳ Подготовка данных для экспорта...")
    
    try:
        if export_format == "excel":
            filename = export_data_to_excel(data_type, start_date, end_date)
            
            with open(filename, 'rb') as file:
                await bot.send_document(
                    callback_query.from_user.id,
                    types.InputFile(file, filename=f"export_{data_type}_{start_date}_to_{end_date}.xlsx"),
                    caption=f"Экспорт данных ({data_type}) с {start_date} по {end_date}"
                )
        else:  # JSON
            filename = export_data_to_json(data_type, start_date, end_date)
            
            with open(filename, 'rb') as file:
                await bot.send_document(
                    callback_query.from_user.id,
                    types.InputFile(file, filename=f"export_{data_type}_{start_date}_to_{end_date}.json"),
                    caption=f"Экспорт данных ({data_type}) с {start_date} по {end_date}"
                )
        
        # Clean up temp file
        try:
            os.remove(filename)
        except:
            pass
        
        # Reset state
        await state.finish()
        
        # Show main menu
        keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(KeyboardButton("📊 Статистика"))
        keyboard.add(KeyboardButton("🔍 Поиск контента"))
        keyboard.add(KeyboardButton("📤 Экспорт данных"))
        keyboard.add(KeyboardButton("📋 Управление источниками"))
        keyboard.add(KeyboardButton("🔑 Ключевые слова"))
        
        await bot.send_message(
            callback_query.from_user.id,
            "✅ Экспорт завершен. Чем еще могу помочь?",
            reply_markup=keyboard
        )
    
    except Exception as e:
        logger.error(f"Error during export: {e}")
        await bot.send_message(
            callback_query.from_user.id,
            f"❌ Ошибка при экспорте данных: {e}"
        )
        await state.finish()

@dp.callback_query_handler(lambda c: c.data == "back_to_main", state="*")
async def back_to_main(callback_query: types.CallbackQuery, state: FSMContext):
    """Return to main menu"""
    await callback_query.answer()
    await state.finish()
    
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(KeyboardButton("📊 Статистика"))
    keyboard.add(KeyboardButton("🔍 Поиск контента"))
    keyboard.add(KeyboardButton("📤 Экспорт данных"))
    keyboard.add(KeyboardButton("📋 Управление источниками"))
    keyboard.add(KeyboardButton("🔑 Ключевые слова"))
    
    await callback_query.message.answer("Главное меню:", reply_markup=keyboard)
    await callback_query.message.delete()

@dp.callback_query_handler(lambda c: c.data == "back_to_period", state=ExportStates.select_format)
async def back_to_period_selection(callback_query: types.CallbackQuery, state: FSMContext):
    """Return to period selection"""
    await callback_query.answer()
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("Неделя", callback_data="period_week"))
    keyboard.add(InlineKeyboardButton("Месяц", callback_data="period_month"))
    keyboard.add(InlineKeyboardButton("3 месяца", callback_data="period_three_months"))
    keyboard.add(InlineKeyboardButton("Все время", callback_data="period_all"))
    keyboard.add(InlineKeyboardButton("Свой период", callback_data="period_custom"))
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_export_type"))
    
    await callback_query.message.edit_text("Выберите период:", reply_markup=keyboard)
    await ExportStates.select_period.set()

@dp.callback_query_handler(lambda c: c.data == "back_to_export_type", state=ExportStates.select_period)
async def back_to_export_type_selection(callback_query: types.CallbackQuery, state: FSMContext):
    """Return to export type selection"""
    await callback_query.answer()
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("Посты", callback_data="export_posts"))
    keyboard.add(InlineKeyboardButton("Комментарии", callback_data="export_comments"))
    keyboard.add(InlineKeyboardButton("Сообщения из групп", callback_data="export_messages"))
    keyboard.add(InlineKeyboardButton("Все данные", callback_data="export_all"))
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_main"))
    
    await callback_query.message.edit_text("Выберите тип данных для экспорта:", reply_markup=keyboard)
    await ExportStates.select_data_type.set()

# Source management handlers
@dp.message_handler(lambda message: message.text == "📋 Управление источниками")
async def manage_sources_command(message: types.Message):
    """Show source management options"""
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("➕ Добавить источник", callback_data="add_source"))
    keyboard.add(InlineKeyboardButton("📃 Список источников", callback_data="list_sources"))
    keyboard.add(InlineKeyboardButton("❌ Удалить источник", callback_data="delete_source"))
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_main"))
    
    await message.answer("Управление источниками:", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data == "add_source")
async def add_source_command(callback_query: types.CallbackQuery):
    """Start add source flow"""
    await callback_query.answer()
    
    await callback_query.message.edit_text(
        "Введите имя источника (имя канала или группы без @):"
    )
    await SourceStates.add_source.set()

@dp.message_handler(state=SourceStates.add_source)
async def process_source_name(message: types.Message, state: FSMContext):
    """Process source name input"""
    source_name = message.text.strip()
    
    if source_name.startswith('@'):
        source_name = source_name[1:]
    
    await state.update_data(source_name=source_name)
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("Канал", callback_data="source_type_channel"))
    keyboard.add(InlineKeyboardButton("Группа", callback_data="source_type_group"))
    
    await message.answer(f"Выберите тип источника для '{source_name}':", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data.startswith('source_type_'), state=SourceStates.add_source)
async def process_source_type(callback_query: types.CallbackQuery, state: FSMContext):
    """Process source type selection"""
    await callback_query.answer()
    
    source_type = callback_query.data.split('_')[2]
    data = await state.get_data()
    source_name = data.get('source_name')
    
    result = add_source(source_name, source_type)
    
    if result:
        await callback_query.message.edit_text(f"✅ Источник '{source_name}' успешно добавлен!")
    else:
        await callback_query.message.edit_text(f"❌ Источник '{source_name}' уже существует или произошла ошибка.")
    
    await state.finish()
    
    # Show sources management menu after delay
    await asyncio.sleep(2)
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("➕ Добавить источник", callback_data="add_source"))
    keyboard.add(InlineKeyboardButton("📃 Список источников", callback_data="list_sources"))
    keyboard.add(InlineKeyboardButton("❌ Удалить источник", callback_data="delete_source"))
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_main"))
    
    await callback_query.message.edit_text("Управление источниками:", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data == "list_sources")
async def list_sources_command(callback_query: types.CallbackQuery):
    """List all monitored sources"""
    await callback_query.answer()
    
    sources = get_sources()
    
    if not sources:
        await callback_query.message.edit_text(
            "📂 Список источников пуст.\n\n"
            "Нажмите '➕ Добавить источник' для добавления нового источника.",
            reply_markup=InlineKeyboardMarkup().add(
                InlineKeyboardButton("« Назад", callback_data="back_to_sources")
            )
        )
        return
    
    sources_text = "📂 Список отслеживаемых источников:\n\n"
    
    for i, (name, type_) in enumerate(sources, 1):
        source_type = "Канал" if type_ == "channel" else "Группа"
        sources_text += f"{i}. {name} - {source_type}\n"
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_sources"))
    
    await callback_query.message.edit_text(sources_text, reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data == "delete_source")
async def delete_source_command(callback_query: types.CallbackQuery):
    """Start delete source flow"""
    await callback_query.answer()
    
    sources = get_sources()
    
    if not sources:
        await callback_query.message.edit_text(
            "📂 Список источников пуст.\n\n"
            "Нет источников для удаления.",
            reply_markup=InlineKeyboardMarkup().add(
                InlineKeyboardButton("« Назад", callback_data="back_to_sources")
            )
        )
        return
    
    keyboard = InlineKeyboardMarkup(row_width=1)
    
    for name, _ in sources:
        keyboard.add(InlineKeyboardButton(name, callback_data=f"delete_{name}"))
    
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_sources"))
    
    await callback_query.message.edit_text("Выберите источник для удаления:", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data.startswith('delete_'))
async def confirm_delete_source(callback_query: types.CallbackQuery, state: FSMContext):
    """Confirm source deletion"""
    await callback_query.answer()
    
    source_name = callback_query.data[7:]  # Remove 'delete_' prefix
    await state.update_data(source_name=source_name)
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("✅ Да", callback_data="confirm_delete_yes"))
    keyboard.add(InlineKeyboardButton("❌ Нет", callback_data="confirm_delete_no"))
    
    await callback_query.message.edit_text(
        f"Вы уверены, что хотите удалить источник '{source_name}'?",
        reply_markup=keyboard
    )
    await SourceStates.confirm_delete.set()

@dp.callback_query_handler(lambda c: c.data.startswith('confirm_delete_'), state=SourceStates.confirm_delete)
async def process_delete_confirmation(callback_query: types.CallbackQuery, state: FSMContext):
    """Process delete confirmation"""
    await callback_query.answer()
    
    data = await state.get_data()
    source_name = data.get('source_name')
    
    if callback_query.data == "confirm_delete_yes":
        delete_source(source_name)
        await callback_query.message.edit_text(f"✅ Источник '{source_name}' удален.")
    else:
        await callback_query.message.edit_text("❌ Удаление отменено.")
    
    await state.finish()
    
    # Show sources management menu after delay
    await asyncio.sleep(2)
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("➕ Добавить источник", callback_data="add_source"))
    keyboard.add(InlineKeyboardButton("📃 Список источников", callback_data="list_sources"))
    keyboard.add(InlineKeyboardButton("❌ Удалить источник", callback_data="delete_source"))
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_main"))
    
    await callback_query.message.edit_text("Управление источниками:", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data == "back_to_sources")
async def back_to_sources_menu(callback_query: types.CallbackQuery):
    """Return to sources management"""
    await callback_query.answer()
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("➕ Добавить источник", callback_data="add_source"))
    keyboard.add(InlineKeyboardButton("📃 Список источников", callback_data="list_sources"))
    keyboard.add(InlineKeyboardButton("❌ Удалить источник", callback_data="delete_source"))
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_main"))
    
    await callback_query.message.edit_text("Управление источниками:", reply_markup=keyboard)

# Keywords management handlers
@dp.message_handler(lambda message: message.text == "🔑 Ключевые слова")
async def manage_keywords_command(message: types.Message):
    """Show keywords management options"""
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("➕ Добавить ключевое слово", callback_data="add_keyword"))
    keyboard.add(InlineKeyboardButton("📃 Список ключевых слов", callback_data="list_keywords"))
    keyboard.add(InlineKeyboardButton("❌ Удалить ключевое слово", callback_data="delete_keyword"))
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_main"))
    
    await message.answer("Управление ключевыми словами:", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data == "add_keyword")
async def add_keyword_command(callback_query: types.CallbackQuery):
    """Start add keyword flow"""
    await callback_query.answer()
    
    await callback_query.message.edit_text(
        "Введите ключевое слово для отслеживания:"
    )
    await KeywordStates.add_keyword.set()

@dp.message_handler(state=KeywordStates.add_keyword)
async def process_keyword(message: types.Message, state: FSMContext):
    """Process keyword input"""
    keyword = message.text.strip().lower()
    
    result = add_keyword(keyword)
    
    if result:
        await message.answer(f"✅ Ключевое слово '{keyword}' успешно добавлено!")
    else:
        await message.answer(f"❌ Ключевое слово '{keyword}' уже существует или произошла ошибка.")
    
    await state.finish()
    
    # Show keywords management menu
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("➕ Добавить ключевое слово", callback_data="add_keyword"))
    keyboard.add(InlineKeyboardButton("📃 Список ключевых слов", callback_data="list_keywords"))
    keyboard.add(InlineKeyboardButton("❌ Удалить ключевое слово", callback_data="delete_keyword"))
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_main"))
    
    await message.answer("Управление ключевыми словами:", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data == "list_keywords")
async def list_keywords_command(callback_query: types.CallbackQuery):
    """List all monitored keywords"""
    await callback_query.answer()
    
    keywords = get_keywords()
    
    if not keywords:
        await callback_query.message.edit_text(
            "🔑 Список ключевых слов пуст.\n\n"
            "Нажмите '➕ Добавить ключевое слово' для добавления нового ключевого слова.",
            reply_markup=InlineKeyboardMarkup().add(
                InlineKeyboardButton("« Назад", callback_data="back_to_keywords")
            )
        )
        return
    
    keywords_text = "🔑 Список отслеживаемых ключевых слов:\n\n"
    
    for i, keyword in enumerate(keywords, 1):
        keywords_text += f"{i}. {keyword}\n"
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_keywords"))
    
    await callback_query.message.edit_text(keywords_text, reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data == "delete_keyword")
async def delete_keyword_command(callback_query: types.CallbackQuery):
    """Start delete keyword flow"""
    await callback_query.answer()
    
    keywords = get_keywords()
    
    if not keywords:
        await callback_query.message.edit_text(
            "🔑 Список ключевых слов пуст.\n\n"
            "Нет ключевых слов для удаления.",
            reply_markup=InlineKeyboardMarkup().add(
                InlineKeyboardButton("« Назад", callback_data="back_to_keywords")
            )
        )
        return
    
    keyboard = InlineKeyboardMarkup(row_width=1)
    
    for keyword in keywords:
        keyboard.add(InlineKeyboardButton(keyword, callback_data=f"delete_kw_{keyword}"))
    
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_keywords"))
    
    await callback_query.message.edit_text("Выберите ключевое слово для удаления:", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data.startswith('delete_kw_'))
async def confirm_delete_keyword(callback_query: types.CallbackQuery, state: FSMContext):
    """Confirm keyword deletion"""
    await callback_query.answer()
    
    keyword = callback_query.data[10:]  # Remove 'delete_kw_' prefix
    await state.update_data(keyword=keyword)
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("✅ Да", callback_data="confirm_kw_delete_yes"))
    keyboard.add(InlineKeyboardButton("❌ Нет", callback_data="confirm_kw_delete_no"))
    
    await callback_query.message.edit_text(
        f"Вы уверены, что хотите удалить ключевое слово '{keyword}'?",
        reply_markup=keyboard
    )
    await KeywordStates.confirm_delete.set()

@dp.callback_query_handler(lambda c: c.data.startswith('confirm_kw_delete_'), state=KeywordStates.confirm_delete)
async def process_delete_keyword_confirmation(callback_query: types.CallbackQuery, state: FSMContext):
    """Process delete keyword confirmation"""
    await callback_query.answer()
    
    data = await state.get_data()
    keyword = data.get('keyword')
    
    if callback_query.data == "confirm_kw_delete_yes":
        delete_keyword(keyword)
        await callback_query.message.edit_text(f"✅ Ключевое слово '{keyword}' удалено.")
    else:
        await callback_query.message.edit_text("❌ Удаление отменено.")
    
    await state.finish()
    
    # Show keywords management menu after delay
    await asyncio.sleep(2)
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("➕ Добавить ключевое слово", callback_data="add_keyword"))
    keyboard.add(InlineKeyboardButton("📃 Список ключевых слов", callback_data="list_keywords"))
    keyboard.add(InlineKeyboardButton("❌ Удалить ключевое слово", callback_data="delete_keyword"))
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_main"))
    
    await callback_query.message.edit_text("Управление ключевыми словами:", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data == "back_to_keywords")
async def back_to_keywords_menu(callback_query: types.CallbackQuery):
    """Return to keywords management"""
    await callback_query.answer()
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("➕ Добавить ключевое слово", callback_data="add_keyword"))
    keyboard.add(InlineKeyboardButton("📃 Список ключевых слов", callback_data="list_keywords"))
    keyboard.add(InlineKeyboardButton("❌ Удалить ключевое слово", callback_data="delete_keyword"))
    keyboard.add(InlineKeyboardButton("« Назад", callback_data="back_to_main"))
    
    await callback_query.message.edit_text("Управление ключевыми словами:", reply_markup=keyboard)

# Search content handlers
@dp.message_handler(lambda message: message.text == "🔍 Поиск контента")
async def search_content_command(message: types.Message):
    """Start search content flow"""
    await message.answer("Введите поисковый запрос:")
    await SearchStates.enter_query.set()

@dp.message_handler(state=SearchStates.enter_query)
async def process_search_query(message: types.Message, state: FSMContext):
    """Process search query"""
    query = message.text.strip()
    
    if not query:
        await message.answer("Поисковый запрос не может быть пустым. Пожалуйста, введите поисковый запрос:")
        return
    
    await state.update_data(query=query)
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("Неделя", callback_data="search_period_week"))
    keyboard.add(InlineKeyboardButton("Месяц", callback_data="search_period_month"))
    keyboard.add(InlineKeyboardButton("3 месяца", callback_data="search_period_three_months"))
    keyboard.add(InlineKeyboardButton("Все время", callback_data="search_period_all"))
    keyboard.add(InlineKeyboardButton("Свой период", callback_data="search_period_custom"))
    
    await message.answer(f"Выберите период для поиска '{query}':", reply_markup=keyboard)
    await SearchStates.select_period.set()

@dp.callback_query_handler(lambda c: c.data.startswith('search_period_'), state=SearchStates.select_period)
async def process_search_period(callback_query: types.CallbackQuery, state: FSMContext):
    """Process selected period for search"""
    await callback_query.answer()
    
    period = callback_query.data.split('_')[2]
    
    if period == "custom":
        await callback_query.message.edit_text("Введите начальную дату в формате ГГГГ-ММ-ДД:")
        await SearchStates.custom_period_start.set()
    else:
        # Get start and end dates based on period
        start_date, end_date = get_period_dates(period)
        
        # Get search query
        data = await state.get_data()
        query = data.get('query')
        
        await callback_query.message.edit_text("🔍 Выполняется поиск, пожалуйста, подождите...")
        
        # Perform search
        results = search_content(query, start_date, end_date)
        
        if not results:
            await callback_query.message.edit_text(
                f"❌ По запросу '{query}' ничего не найдено.",
                reply_markup=InlineKeyboardMarkup().add(
                    InlineKeyboardButton("🔙 Новый поиск", callback_data="new_search")
                )
            )
        else:
            # Format results
            result_text = f"🔍 Результаты поиска по запросу '{query}':\n\n"
            
            for i, (date, source, content, content_type) in enumerate(results[:15], 1):  # Limit to 15 results
                formatted_date = date.split()[0] if ' ' in date else date
                result_text += f"{i}. [{formatted_date}] {source} ({content_type}):\n{content[:100]}...\n\n"
            
            if len(results) > 15:
                result_text += f"\nПоказаны первые 15 из {len(results)} результатов."
            
            # Split message if it's too long
            if len(result_text) > 4000:
                chunks = [result_text[i:i+4000] for i in range(0, len(result_text), 4000)]
                for chunk in chunks:
                    await bot.send_message(callback_query.from_user.id, chunk
